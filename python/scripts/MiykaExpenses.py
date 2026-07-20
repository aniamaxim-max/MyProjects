import os
import sys
import pandas as pd

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from config.config import OUTPUT_FILE_MIYKA
from utils.db import query_to_df
from utils.email_sender import send_email


COLUMN_MAP = {
    "OrderNumber": "Номер ЗТС",
    "OrderDate": "Дата розвантаження",
    "Client": "Замовник",
    "Manager": "Менеджер",
    "ReasonFilter": "Причина",
    "DocCount": "К-сть документів мийки",
    "TotalWithoutVAT_EUR": "Сума без ПДВ, євро",
    "TotalSum_EUR": "Сума з ПДВ, євро",
    "TotalVAT_EUR": "ПДВ, євро",
}


def export_and_send():
    query = "EXEC pbi.GetMiykaExpenses"

    print("Підключення до бази даних...")
    try:
        df = query_to_df(query)
        print(f"Запит успішно виконано. Отримано рядків: {len(df)}")

        df.rename(columns=COLUMN_MAP, inplace=True)

        # Excel з форматуванням
        with pd.ExcelWriter(OUTPUT_FILE_MIYKA, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Звіт")
            ws = writer.sheets["Звіт"]

            ws.freeze_panes = "A2"

            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align

            # Формат дати для колонки "Дата розвантаження" (стовпець B)
            for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row):
                for cell in row:
                    cell.number_format = "DD.MM.YYYY"

            # Ширина колонок
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20

        print(f"Дані збережено у файл: {OUTPUT_FILE_MIYKA}")
    except Exception as e:
        print(f"Помилка при роботі з БД: {e}")
        return

    # Email: тільки перші 10 рядків
    table_preview = df.head(10).to_html(index=False, border=0, classes="data-table")

    html_body = f"""
    <html><head>
    <style>
        .data-table {{ border-collapse: collapse; font-family: Arial, sans-serif; font-size: 13px; width: 100%; }}
        .data-table th {{ background-color: #4472C4; color: white; padding: 8px 12px; text-align: left; }}
        .data-table td {{ padding: 6px 12px; border-bottom: 1px solid #ddd; }}
        .data-table tr:nth-child(even) {{ background-color: #f2f2f2; }}
    </style>
    </head><body>
    <h3>Вітаю!</h3>
    <p>У вкладенні знаходиться свіжа вигрузка по запиту <b>Витрати на мийку</b>.</p>
    <p>Попередній перегляд (перші 10 рядків з {len(df)}):</p>
    {table_preview}
    <p>Повний звіт — у файлі Excel у вкладенні.</p>
    </body></html>
    """

    try:
        send_email(
            to="hanna.maksymova@stellar.ua, tetyana.synyava@stellar.ua",
            subject="Автоматична вигрузка: Витрати на мийку",
            body=html_body,
            attachments=[OUTPUT_FILE_MIYKA],
        )
        print("Лист успішно відправлено!")
    except Exception as e:
        print(f"Помилка при відправці пошти: {e}")

    finally:
        if os.path.exists(OUTPUT_FILE_MIYKA):
            os.remove(OUTPUT_FILE_MIYKA)
            print("Тимчасовий файл видалено.")


if __name__ == "__main__":
    export_and_send()
